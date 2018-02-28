#Determines whether proposed condition is in compliance with hydromod by comparing to existing condition.

#Calculate partial duration analysis for San Diego hydromod

# San Diego HMP Manual Ch 6, pg 6-25

# must have 24-hr dry between events
# dry = no rain greater than 0.01 cfs  --> check this against SDHM - cutoff appears to be somewhat arbitrary.
# return a dataset with peak Qs for each event, and the number of peaks.

# ASSUMPTIONS: Q data is continuous, hourly data with no gaps.

import os
import openpyxl
from openpyxl import Workbook
from Flow_freq import Qprocess   #does flow frequency analysis

path = r'H:\pdata\160817\Calcs\Strmwater\Water Quality\EPA SWMM 5\Add DMA1-2_totalarea=9.43ac'

#Existing condition Q out file from EPA SWMM 5
EXdata = 'PRE_DEV_POC-1_QOUT.TXT'

#Proposed condition Q out file from EPA SWMM 5
PRdata = 'POST_DEV_POC-1_QOUT.TXT'

os.chdir(path)

#generate existing flow values in 0.1*Q2 - Q10 to compare to:
EXoutputs = Qprocess(path,EXdata)
EXQcompare = EXoutputs[0]    #comparison bins
EXQ =  EXoutputs[1]

#get list of proposed condition peaks
PRoutputs = Qprocess(path,PRdata)
#PRQcompare = PRoutputs[0]    #not needed for hydromod compliance determination
PRQ = PRoutputs[1]

#2. count number of hours in the flow record that Q exceeds each of the 100 Qs. Do for existing Qs in this script --> sort and then find first exceedance; subtract.

#for each "bin" in existing condition Qcompare, count number of existing and number of proposed that exceed.
EXcomply = []
PRcomply = []
Qfractionlist = []

for Qbin in EXQcompare:
    EXcount = 0
    PRcount = 0

    for qqex in EXQ:
        if qqex > Qbin:
            EXcount +=1
    for qqpr in PRQ:
        if qqpr > Qbin:
            PRcount +=1
    PRcomply.append(PRcount)
    EXcomply.append(EXcount)

    Qfraction = 1.0*PRcount/(1.0*EXcount)
    Qfractionlist.append(Qfraction)

    if Qfraction > 1.1:
        print "FAIL: PR/EX:", Qfraction


#print("pr", PRcomply)
#print("/n" "ex", EXcomply)


#write peak lists to file for comparison using openpyxl --> done

wb = Workbook()
ws = wb.active

#set column titles
ws.cell(row=1, column=1).value="Qbin - existing"
ws.cell(row=1, column=2).value="Existing hours exceeding"
ws.cell(row=1, column=3).value="Proposed hours exceeding"
ws.cell(row=1, column=4).value="PR/EX"
ws.cell(row=1, column=5).value="PASS/FAIL"

i=0
for Qbin in EXQcompare:
    ws.cell(row=i+2, column=1).value = EXQcompare[i]
    ws.cell(row=i+2, column=2).value = EXcomply[i]
    ws.cell(row=i+2, column=3).value = PRcomply[i]
    ws.cell(row=i+2, column=4).value = Qfractionlist[i]
    if Qfraction > 1.1:
        ws.cell(row=i+2, column=5).value = "FAIL"
    else:
        ws.cell(row=i+2, column=5).value = "PASS"

    i+=1

wb.save('Hydromod_compliance'+EXdata+'.xlsx')

