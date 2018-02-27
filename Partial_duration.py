#Processes precip gage depth vs time data.


#technically, can look at TW list of peaks for the occurrence date, and that is the depth corresponding to that year.
#can back calc using existing condition.


#Calculate partial duration analysis for San Diego hydromod

# San Diego HMP Manual Ch 6, pg 6-25

# must have 24-hr dry between events
# dry = no rain greater than 0.01 cfs  --> check this against SDHM - cutoff appears to be somewhat arbitrary.
# return a dataset with peak Qs for each event, and the number of peaks.

# ASSUMPTIONS: Q data is continuous, hourly data with no gaps.


# NOTE: Tory Walkers did this differently from the manual - they got the peak within a 25hr period.

import os
import csv
import openpyxl
from openpyxl import Workbook

path = r'H:\pdata\160817\Calcs\Strmwater\Water Quality\Compare gage data'
indata = 'Oceanside_gage_SDHM.csv'
low = 0.01
sep = 24  #24hr between events

date = []
Q = []
Qpart = []
dry = 24  #number of dry hours - start at 24 to catch first event, which can occur at hr = 0
peak = 0
peaklist = []

os.chdir(path)

with open(indata, 'r') as f:
    csvreader = csv.reader(f, delimiter = ',')
    for row in csvreader:
        try:
            date.append(row[0])
            Q.append(float(row[2]))
        except ValueError:
            #print("can't convert to float:", row)
            pass

series = list(zip(date, Q))  #just need Qs if data is hourly, but leave for now.
#print ("series", series)

#HMP manual way of determining series:

for event in series:
    if event[1] <= low:
        dry+= 1   #Count dry hour
    if event[1] > low:
        #print("rain", event[1])

        if dry >= 24:  #new event

            if peak > 0:
                peaklist.append(peak)  # append peak from last event
                #print("peak appended", peak)
                peak = 0    #reset peak for current storm
        #else:   #same event because less than 24 dry hours have passed
        if event[1] > peak:
            peak = event[1]
            #print("current peak", peak)

        dry = 0  #reset dry hours

#print("len list", len(peaklist))

#peaklist.append(peak)  #peak from last event
#del(peaklist[0])
print("peaklist", peaklist)
#print("HMP peak:", peaklist)


#Tory walkers way of determining peaks:
def TWmethod(series, low):
    peakTW = 0
    peaklistTW = []
    ii = 0

    for ii in range(0, len(series)):
        #print (series[ii][1])


        if series[ii][1] > peakTW and series[ii][1] > low:
            peakTW = series[ii][1]
            #print("test peak:", peakTW)

            #check if next 12 values are decreasing by getting their max and comparing to current peak
            if series[ii + 1][1] < series[ii][1]:
                maxnext = max([zz[1] for zz in series[1+ii : 12+ii]])
                if maxnext < peakTW:
                    peaklistTW.append(peakTW)
                    peakTW = 0
        ii += 1

    print "\n" "Tory Walker Peaks", peaklistTW

    return(peaklistTW)

#write peak lists to file for comparison using openpyxl
wb = Workbook()
ws = wb.active

#set column titles
ws.cell(row=1, column=1).value="San Diego Manual: 24hr dry btwn storms"
ws.cell(row=1, column=4).value="Tory Walker: 12hr buffer after peak"

ws.cell(row=2, column=1).value="Date"
ws.cell(row=2, column=2).value="Depth, in"

ws.cell(row=2, column=4).value="Date"
ws.cell(row=2, column=5).value="Depth, in"


peaklistTW = TWmethod(series, low)

i=3
for item in peaklist:
    ws.cell(row=i, column=2).value = item
    i+=1

i=3
for item in peaklistTW:
    ws.cell(row=i, column=5).value = item
    i+=1


wb.save('Compare_peak.xlsx')


"""
#Calculate flow frequencies for San Diego hydromod

# San Diego HMP Manual Ch 6, pg 6-25

#1. sort values large to small
#2. Flow frequency with Cunnane Eqn: Probability = P = (i-0.4)/(n+0.2)
#   i = Position of the peak whose probability is desired
#   n = number of years analyzed = 1-10 = 10

# ASSUMPTIONS: number of peaks analyzed MUST be less than or equal to n.

#sort Q list


def break1(list, sorted):
    num = len(list)
    mid = int(num/2)

    #print("list", list, "len", num, "mid", mid)
    if mid >=1:
        left = list[:mid]
        right = list[mid:]
        #print("L", left, "R", right)

        break1(left, sorted)
        break1(right, sorted)

    else:
       return sort1(list, sorted)



#sort largest to smallest
def sort1(list, sorted):

    if list[0] >= sorted[0]:
        sorted.insert(0, list[0])

    if list[0] < sorted[0]:
        sorted.append(list[0])

    return sorted

#initialize and call:
sorted = [0]
#list = [1,2,3,4,5,6,7,8,9,9, 9.2]
break1(peaklist, sorted)

#delete last element - it is a zero I had to add for comparison.
#print("FINAL RESULT", sorted[:-1])


#Calculate flow frequencies for San Diego hydromod

# San Diego HMP Manual Ch 6, pg 6-25

#2. Flow frequency with Cunnane Eqn: Probability = P = (i-0.4)/(n+0.2)
#   i = Position of the peak whose probability is desired
#   n = number of years analyzed = 57 (years in flow record)

# ASSUMPTIONS: number of peaks analyzed MUST be equal to n.


n = 57 #years in flow record

for i in range(0, 57):
    #rank = i + 1
    P = ((i+1)-0.4) / (n+0.2)
    #print (P)

#QUESTIONS:
#       will I get 57 peaks from flow record? probably not. how to assign numbers if not right number of peaks?
#NEXT: zip the 57 peaks to P1-P57. Interpolate to get Q1-Q10 (though TW didn't use linear interpolation, judging by Q10)
#1. divide flows of interest (0.1 * Q2 --> Q10) into 100 equal intervals
#2. count number of hours that Q exceeds each of the 100 Qs. Do for existing Qs in this script --> sort and then find first exceedance; subtract.
#2. for proposed, use a different script: sort and count.


"""

