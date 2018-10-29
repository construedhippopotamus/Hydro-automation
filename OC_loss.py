# -*- coding: utf-8 -*-
"""
Orange County loss rate calculation with pandas

Python 3.6

"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
#NOTE: requires xlrd

#read in csv
cn_csv = pd.read_csv('cn.csv', index_col=None)
nat_cn = pd.DataFrame(cn_csv)

#input data
#iterate through sheets
#sheet = 'Sheet1'  #change this to first worksheet - easy in openpyxl; how to do with pandas df???
#read xlsx input
test_input = pd.read_excel('inputs.xlsx', 1)
"""working with multiple sheets: http://pandas.pydata.org/pandas-docs/stable/io.html#excelfile-class  """

#input_data = pd.DataFrame(pd.read_csv('inputs.csv'))
#print("\n", input_data)

#function to return CN for soil, cover in input sheet
def match_cn(input_df, nat_cn):
    
    res_cn_list = []
    res_ap_list = []
    
    for ii in range(len(input_df)):
        soil_in = input_df.Soil[ii]
        cover_in = input_df.Cover[ii]
        
        result_cn = nat_cn.loc[nat_cn.Cover == cover_in][soil_in].item()
        result_ap = nat_cn.ap.loc[nat_cn.Cover == cover_in].item()
        
        #print(ii, result_cn, result_ap)
        res_cn_list.append(result_cn)
        res_ap_list.append(result_ap)
        
    return res_cn_list, res_ap_list

#fp is the maximum loss rates, which occurs in 100% pervious cover
fp_dict = {'A': 0.4, 'B': 0.3, 'C': 0.25, 'D': 0.20}
fp_lookup = pd.Series(fp_dict)
#print(fp_df['A'])

"""--------------------------------------------------------------------------"""

#p24 = 24-hr precip depth, inches
p24 = 5.6

#rain intensity in in/hr
intensity = 1.9

#curve number as a function of soil and cover, ap as function of cover
#CN, ap = match_cn(input_data, nat_cn)
cn_list, ap_list = match_cn(input_data, nat_cn)

#area array
area_list = input_data['Area']

#pervious area loss rates
fp_list = (fp_lookup[input_data.Soil]).tolist()
#print('fp', fp) 

#loss rate calculations
def loss_calc(cn_list, p24, intensity, ap_list, fp_list, area_list):

    #yj in list
    yield_list = []
    #fm for each row in list
    fm_list = []
    
    #error checking before calculation
    assert len(cn_list)==len(ap_list)==len(fp_list)==len(area_list), "ERROR: Input data is different length"
    
    #iterate through rows in input data. Calculate yj and fm per row.
    for ii in range(len(cn_list)):
        #storage in soil, s
        s = 1000/cn_list[ii] - 10
        
        # Initial Abstraction,ia
        ia = 0.2*s
      
        #24-hr runoff yield fraction for each CN
        # In OC, other storm durations not supported.
        if ia > p24:
            yj = 0
        else:
            yj = (p24 - ia)**2 / (p24 - ia + s)*p24
        
        #append yj to yield array
        yield_list.append(yj)
        #print(yj)
        
        #fm = max loss rate, calculation of actual loss rates: accounts for imperviousness
        fm = ap_list[ii]*fp_list[ii]
        
        fm_list.append(fm)

    
    #change lists to numpy arrays
    yield_arr = np.array(yield_list)
    area_arr = np.array(area_list)
    fm_arr = np.array(fm_list)
    ap_arr = np.array(ap_list)
    
    area_sum = area_arr.sum()
    
    #area-weighted overall CN, 24-hr
    y = (np.multiply(yield_arr, area_arr)).sum() / area_sum
    
    #low loss fraction, ybar
    ybar = 1- y
    
    #low loss rate, f_star (in/hr)
    f_star = ybar*intensity
    
    #area-weighted fm, in/hr
    fm_overall = (np.multiply(fm_arr, area_arr)).sum() / area_sum
    
    #area-weighted ap
    ap_overall = (np.multiply(ap_arr, area_arr)).sum() / area_sum

    return f_star, ybar, fm_overall, ap_overall, yield_list, fm_list, area_sum

f_star, ybar, fm_overall, ap_overall, yield_list, fm_list, area_sum = loss_calc(cn_list, p24, intensity, ap_list, fp_list, area_list)
#print("\n", "results")
#print('fm', fm, 'ybar', ybar, 'fstar', f_star)
#print(f_star, ybar, fm_overall)


df = (pd.DataFrame([yield_list, fm_list,  ap_list, fp_list])).transpose()  #, columns = ['yield_arr', 'fm_arr', 'ap_list', 'fp_list']
df.columns = ['yield_arr', 'fm_arr', 'ap_list', 'fp_list']

#append calculation results to original input data
combined = input_data.join(df)

#use same name as input data sheet
combined.to_excel('results.xlsx', sheet_name=sheet)  #this doesn't check for existing workbooks

wb = openpyxl.load_workbook(filename = 'results.xlsx')

ws = wb[sheet] #name from input data tab.

#Summary data
ws['K1'] = 'Summary'
ws['K2'] = 'Ap'
ws['K3'] = 'Fm'
ws['K4'] = 'ybar'
ws['K5'] = 'Area'

ws['J2'] = ap_overall
ws['J3'] = fm_overall
ws['J4'] = ybar
ws['J5'] = area_sum


# Save the file
wb.save("results.xlsx")
#results saved to csv
#combined.to_csv('results.csv')
"""---------------------------------------------"""

        









"""TO DO
- find loss rate calculation mistake

- excel (one subarea per tab) --> tab name and iteration
- file overwrite protection

- add in urban cn to input data sheet
- make code that will create cn sheet in python so don't need excel
- incorporate AMC/storm year rules

- add input data checking! :) string match, case, numbers
    
- gui?
- make dependencies list
- how to make exe (tho unneeded for this)


"""






"""everything below comment is not used yet"""

#create curve number AMC table
AMC2 = [100, 95, 90, 85, 80, 75, 70, 65, 60, 55, 50, 45, 40, 35, 30, 25, 20, 15, 10, 5, 0]
AMC1 = [100, 87, 78, 70, 63, 57, 51, 45, 40, 35, 31, 27, 23, 19, 15, 12, 9, 7, 4, 2, 0]
AMC3 = [100, 99, 98, 97, 94, 91, 87, 83, 79, 75, 70, 65, 60, 55, 50, 45, 39, 33, 26, 17, 0]

#add interpolation?
#AMC is a function of the storm frequency year.
AMC_table = pd.DataFrame({'AMC2': AMC2, 'AMC1': AMC1, 'AMC3': AMC3})
#print(AMC_table)

#pervious fractions for urban covers
urban_ap = {'Public Park': 85, 'School': 60, '2.5 ac lots': 90, '1 ac lots': 80,
        '2 DU/ac': 70, '3-4DU/ac':60, '5-7DU/ac':50, '8-10DU/ac': 40, '10+DU/ac': 20,
        'Condo': 35, 'Apartment': 20, 'Mobile Home Park': 25, 'Commercial': 10
        }

#curve numbers for urban cover
#urban_cn = {'Landscaping Good': [32, 56, 69, 75]}
soil_names = ['A', 'B', 'C', 'D']

urban_cn_num = np.array([[32, 56, 69, 75], [58, 74, 83, 87], [44, 65, 77, 82], [33, 58, 72, 79]] ).reshape(4,4)
#print(urban_cn)
urban_cover = ['Landscaping Good', 'Turf poor', 'Turf fair', 'Turf good']

# regex to match : ## or :##  :.\d\d  .=any
urban_ap = ['Public Park', 'School', '2.5 ac lots', '1 ac lots',
        '2 DU/ac', '3-4DU/ac', '5-7DU/ac', '8-10DU/ac', '10+DU/ac',
        'Condo', 'Apartment', 'Mobile Home Park', 'Commercial'
        ]

urban_cn = pd.DataFrame(urban_cn_num, columns = soil_names, index=urban_cover)
#print(s)

