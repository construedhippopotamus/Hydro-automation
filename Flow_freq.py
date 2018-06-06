#Calculates return periods for existing condition flows.

#Calculate partial duration analysis for San Diego hydromod

# San Diego HMP Manual Ch 6, pg 6-25

# must have 24-hr dry between events
# dry = no rain greater than 0.01 cfs  --> check this against SDHM - cutoff appears to be somewhat arbitrary.
# return a dataset with peak Qs for each event, and the number of peaks.

# ASSUMPTIONS: Q data is continuous, hourly data with no gaps.


# NOTE: Tory Walkers did this differently from the manual - they got the peak within a 25hr period.

import os
import csv
import openpyxl
from openpyxl import Workbook

#path = r'H:\pdata\160817\Calcs\Strmwater\Water Quality\EPA SWMM 5\Original Tory Walker models'
#Existing condition Q out file from EPA SWMM 5
#EXdata = 'PRE_DEV_POC-1_QOUT.TXT'

def Qprocess(path, datafile):

    low = 0.01
    sep = 24  #24hr between events

    date = []
    Q = []
    Qpart = []
    dry = 24  #number of dry hours - start at 24 to catch first event, which can occur at hr = 0
    peak = 0
    peaklist = []

    os.chdir(path)

    #Extract Qs from both output files.
    with open(datafile, 'r') as f:
        csvreader = csv.reader(f, delimiter = ' ')
        for row in csvreader:
            try:
                date.append(row[12]) #would need rows 1-3
                Q.append(float(row[23]))
            except ValueError:
                pass
            except IndexError:
                pass

    series = list(zip(date, Q))  #just need Qs if data is hourly, but leave for now.
    #print ("series", series)

    #HMP manual way of determining series:

    for event in series:
        if event[1] <= low:
            dry+= 1   #Count dry hour
        if event[1] > low:
            #print("rain", event[1])

            if dry >= 24:  #new event

                if peak > 0:
                    peaklist.append(peak)  # append peak from last event
                    #print("peak appended", peak)
                    peak = 0    #reset peak for current storm
            #else:   #same event because less than 24 dry hours have passed
            if event[1] > peak:
                peak = event[1]
                #print("current peak", peak)

            dry = 0  #reset dry hours

    peaklist.append(peak) #add last peak
    #print("peaklist", peaklist)


    #Calculate flow frequencies for San Diego hydromod

    # San Diego HMP Manual Ch 6, pg 6-25

    #1. sort values large to small
    #2. Flow frequency with Cunnane Eqn: Probability = P = (i-0.4)/(n+0.2)
    #   i = Position of the peak whose probability is desired
    #   n = number of years analyzed = 1-10 = 10

    # ASSUMPTIONS: number of peaks analyzed MUST be less than or equal to n.


    #Calculate flow frequencies for San Diego hydromod

    # San Diego HMP Manual Ch 6, pg 6-25

    #2. Flow frequency with Cunnane Eqn: Probability = P = (i-0.4)/(n+0.2)
    #   i = Position of the peak whose probability is desired
    #   n = number of years analyzed = 57 (years in flow record)

    # ASSUMPTIONS: number of peaks analyzed MUST be equal to n.

    #remove any duplicates
    peaklist1 = list(set(peaklist))
    #print("len w/o duplicates:", len(peaklist1), "orig len:", len(peaklist))

    #sort peak list
    peaklist1.sort(reverse = True)
    #print("Sorted peak list", peaklist1)

    #keep 57 largest Qs for analysis.
    del(peaklist1[57:])

    #Calculate Cunnane probability for 57 years of flow record at this gage
    n = 57 #years in flow record
    Returnlist = []   #list to store return periods

    for i in range(0, 57):
        #rank = i + 1
        P = ((i+1)-0.4) / (n+0.2)
        R = 1/ P  # return period
        Returnlist.append(R)

    rankQs = dict(zip(Returnlist, peaklist1))

    #print(datafile, rankQs)

    #write rankQs to file for report using openpyxl

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value="Return period, years"
    ws.cell(row=1, column=2).value="Q, cfs"

    j=1
    for RR in Returnlist:
        j+=1
        ws.cell(row=j, column=1).value = RR

	i=1
    for Qp in peaklist1:
        i+=1
        ws.cell(row=i, column=2).value = Qp


    wb.save(datafile + 'rankQ.xlsx')

    #dictionary isn't in order by default in python 2.7... would have to use ordered dict.

    Q2 =  rankQs.get(2.0)
    Q10up = rankQs.get(10.214285714285715)
    Q10low = rankQs.get(8.666666666666668)

    Q10 = Q10low + (10-8.666)* (Q10up - Q10low) / (10.214285714285715-8.666666666666668)

    #Make list of 100 comparison points from 0.1*Q2 --> Q100
    Qcompare = [(Q10 - 0.1*Q2)*y/99 + 0.1*Q2 for y in range(0, 100)]


    print(datafile, "Q2", Q2, "Q10", Q10)


    return Qcompare, Q

