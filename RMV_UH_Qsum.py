# -*- coding: utf-8 -*-
"""
Created on Sat Jan 06 13:38:42 2018

@author: Jenny

NEW UPDATED VERSION - ONE SUMMARY WKBK WITH TABS FOR EACH YEAR

Program to get Qmax from each AES UH file in specified subfolders within main folder (base). Qmax sorted and summed to get
totals tributary to San Juan Creek (SJC) and Gobernadora (GOB). Creates one summary sheet per base folder.

Things that didn't work: Built-in delimiter in csv reader doesn't work because spacing inconsistent and multi-character.
Data from csv reader is a list. List cannot be modified with split on space because space is part of list.
Regular read() reads data one character at a time.
When importing data in pandas, need to set types etc to avoid NaN error. Too much work.

"""
import os
import openpyxl
from openpyxl import Workbook
import csv

base = r"H:\pdata\134519\Calcs\Strmwater\PA-3 & PA-4\Hydrology\Local\Existing\Unit Hydrograph"

#change to folder names if different
pathlist=['2-YR', '5-YR', '10-YR', '25-YR', '50-YR', '100-YR']

#for single folder:
#path=r"H:\pdata\134519\Calcs\Strmwater\PA-3 & PA-4\Hydrology\Local\Existing\Unit Hydrograph\2-YR"

#make new workbook to save data
wb = Workbook()

#add summary sheet
summary=wb.create_sheet("Summary")
summary.cell(row=2, column=1).value="Year"
summary.cell(row=2, column=2).value="SJC"
summary.cell(row=2, column=3).value="GOB"

#iterator for summary sheet rows
j=3

#path=r"H:\pdata\134519\Calcs\Strmwater\PA-3 & PA-4\Hydrology\Local\Existing\Unit Hydrograph\2-YR"
#for each folder in pathlist

for folder in pathlist:
    path = os.path.join(base, folder)

    os.chdir(path)
    #print "path", path

    ws=wb.create_sheet(folder)

    #set column titles
    ws.cell(row=1, column=1).value="Filename"
    ws.cell(row=1, column=2).value="Max Q, cfs"
    ws.cell(row=1, column=3).value="Watershed"
    ws.cell(row=1, column=5).value="SJC sum"
    ws.cell(row=1, column=6).value="GOB sum"

    #excel workbook row iterator
    i=3

    #make file name list
    filelist=[]
    #print "filelist should be empty", filelist
    #PA3 files
    SJC = ['17A', '17B', '311', '312', '313', '314', '315', '316', '318']
    GOB = ['306', '307', '308', '310', '34A', '34B', '35A', '35B', '35C', '39A', '39B' ]
    SJC_sum = []
    GOB_sum = []
    SJC_files = []
    GOB_files = []

    #print "Gob files should be empty", GOB_files

    allfiles = os.listdir(path)

    #get *.RES files
    for file in allfiles:
        if file.upper().endswith('RES'):
            filelist.append(file)
    #print "filelist2", filelist
    for resfile in filelist:

        with open(resfile, 'rb') as data:

            #reset time series datasets
            vol=[]
            hrs=[]
            Q=[]
            #delimiter purposely set to tab, which is not in file, in order to put each line in a cell.
            #need csv reader --> just using text
            reader=csv.reader(data, delimiter='\t')

            for row in reader:
                str1 = str(row)

                if "Q" in str1:
                    #strip out other characters
                    hrs.append(str1[0:11].strip("[").strip("'").strip("-"))
                    vol.append(str1[12:25].strip("[").strip("'").strip("-"))
                    Q.append(str1[23:32].strip("[").strip("'").strip("-"))

        Qmax=float(max(Q[4:]))
        ws.cell(row=i, column=1).value=resfile
        ws.cell(row=i, column=2).value=Qmax

        #conditions for adding to San Juan Creek or Gobernadora
        if resfile[:3] in SJC:
            SJC_sum.append(float(Qmax))
            SJC_files.append(resfile)
            ws.cell(row=i, column=3).value="SJC"

        elif resfile[:3] in GOB:
            GOB_sum.append(float(Qmax))
            GOB_files.append(resfile)
            ws.cell(row=i, column=3).value="GOB"
        #update excel worksheet row iterator
        i=i+1

    Ssummary=sum(SJC_sum)
    Gsummary=sum(GOB_sum)
    ws.cell(row=2, column=5).value=Ssummary
    ws.cell(row=2, column=6).value=Gsummary

    summary.cell(row=j, column=1).value=folder
    summary.cell(row=j, column=2).value=Ssummary
    summary.cell(row=j, column=3).value=Gsummary

    #update summary sheet iterator
    j=j+1

    wkshtname='Qsummary.xlsx'
    wb.save(os.path.join(base, wkshtname))

print "done"
