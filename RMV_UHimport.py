# -*- coding: utf-8 -*-
"""

@author: Jenny

Single area unit hydrograph or small area hydrograph import to excel. Summary tab and one tab per hydrograph. 

python 2.7

Requires installation of openpyxl

"""
import os
import openpyxl
from openpyxl import Workbook
import csv
import cmd

base = r"C:\Users\Pizzagirl\Documents\programming\UH_extractQ"

#if you want to import from nested folder structure, script will import from foldernames in pathlist:
pathlist=['2-YR', '5-YR', '10-YR', '25-YR', '50-YR', '100-YR']
#pathlist = ['full UH']
#make new workbook to save data
wb = Workbook()

#add summary sheet
summary=wb.create_sheet("Summary")
#summary.cell(row=2, column=1).value="Year"
summary.cell(row=2, column=1).value="File Name"
summary.cell(row=2, column=2).value="Total Vol, ac-ft"
summary.cell(row=2, column=3).value="Q peak, cfs"
summary.cell(row=2, column=4).value="Year"

#iterator for summary sheet rows
j=3

#Ask user if UH should be imported
userinput = raw_input("Print N to only output summary. Hit enter to output summary and hydrographs. ")

userbase = str(raw_input("Paste/type in path to folder holding hydrographs to import. \
                          Hydrographs in any of the following folders will be imported: \
                          ['2-YR', '5-YR', '10-YR', '25-YR', '50-YR', '100-YR']"))

#check if path is valid
if os.path.isdir(userbase):
    base = userbase  #overwrite base in script
else: 
    print("Invalid path entered - using path in script. \n")

#Function to check if excel file exists and if so increment name and recheck
def savename(name, base, counter):
    #make sure we look in file location
    os.chdir(base)
    
    if os.path.isfile(name + str(counter) + '.xlsx'):
        counter += 1
        #name = name + str(counter)
        return savename(name, base, counter)
    else:
        name = name + str(counter)
        print("worksheet name: ", name)
        return name

#main function that looks at .res files for all folders in list above
# and puts all hydrographs in excel (default) and makes a summary table
for folder in pathlist:
    
    path = os.path.join(base, folder)

    if os.path.isdir(path) == True:
        os.chdir(path)
   
    else: # no such folder - skip
        print("Folder {} doesn't exist".format(folder))
        continue

    #excel workbook row iterator
    i=3

    #make file name list
    filelist=[]
 
    allfiles = os.listdir(path)

    #get *.RES files
    for file in allfiles:
        if file.upper().endswith('RES'):
            filelist.append(file)
   
    for resfile in filelist:
        
        if userinput.upper() <> "N":  #add a tab for each hydrograph
            
            #Create one tab per res file
            ws=wb.create_sheet(resfile)
        
            #set column titles
            ws.cell(row=1, column=1).value="Filename"
            ws.cell(row=1, column=2).value="Time, hrs"
            ws.cell(row=1, column=3).value="Vol, ac-ft"
            ws.cell(row=1, column=4).value="Q, cfs"
            ws.cell(row=1, column=5).value="Total vol, ac-ft"            
            ws.cell(row=1, column=6).value="Q max, cfs"

        with open(resfile, 'rb') as data:

            #reset time series datasets
            vol=[]
            hrs=[]
            Q=[]
            #delimiter purposely set to tab, which is not in file, in order to put each line in a cell.
            #need csv reader --> just using text
            reader=csv.reader(data, delimiter='\t')

            for row in reader:
                str1 = str(row)

                if "Q" in str1:
                    #strip out other characters
                    hrs.append(str1[0:11].strip("[").strip("'").strip("-"))
                    vol.append(str1[12:25].strip("[").strip("'").strip("-"))
                    Q.append(str1[23:33].strip("[").strip("'").strip("-"))
        
        #get max values from Q and vol
        Qmax= float(max(Q[4:]))
        Volmax = float(max(vol[4:]))
        
        if userinput.upper() <> "N":  #print one hydrograph per sheet
            
            ws.cell(row=2, column=5).value= Volmax            
            ws.cell(row=2, column=6).value= Qmax
            ws.cell(row=2, column=1).value= resfile
            
            #datasets are shifted to avoid printing out text
            #start writing data at row 3
            for rowiter in range(3, len(hrs)):
                ws.cell(row=rowiter, column=2).value= float(hrs[rowiter])            
                ws.cell(row=rowiter, column=3).value= float(vol[rowiter])  
                ws.cell(row=rowiter, column=4).value= float(Q[rowiter]) 

        #summary tab output
        summary.cell(row=j, column=1).value= resfile
        summary.cell(row=j, column=2).value= Volmax
        summary.cell(row=j, column=3).value= Qmax
        summary.cell(row=j, column=4).value= folder  #year
    
        #update summary sheet iterator
        j=j+1
    
#save excel file
wkshtname='Qsummary'

#try saving with current name, current path. Counter to increment if overwrite
finalwkshtname = savename(wkshtname, base, 1)

wb.save(os.path.join(base, finalwkshtname + ".xlsx"))

print "done"
